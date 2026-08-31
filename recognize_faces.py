import cv2
import pickle
import numpy as np
from deepface import DeepFace
from datetime import datetime
import sqlite3
from openpyxl import Workbook, load_workbook
import os

# ========== DATE ==========
today_date = datetime.now().strftime("%Y-%m-%d")
file_name = f"{today_date}.xlsx"

# ========== DATABASE ==========
conn = sqlite3.connect("attendance.db")
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS attendance (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    date TEXT,
    time TEXT,
    status TEXT
)
""")
conn.commit()

# ========== EXCEL SETUP ==========
if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Date", "Time", "Status"])
    wb.save(file_name)

# ========== LOAD ENCODINGS ==========
with open("encodings.pkl", "rb") as f:
    data = pickle.load(f)

if isinstance(data, dict):
    known_encodings = data["encodings"]
    known_names = data["names"]
else:
    known_encodings, known_names = data

print("Encodings loaded successfully!")

# 🔥 UNIQUE STUDENT LIST
all_students = set(known_names)

# ========== FACE DETECTOR ==========
face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

def cosine_similarity(a, b):
    return np.dot(a, b) / (np.linalg.norm(a) * np.linalg.norm(b))

# ========== CAMERA ==========
video_capture = cv2.VideoCapture(0)
print("Camera started...")

frame_count = 0
process_every_n_frames = 10

present_today = set()
face_results = {}

while True:
    ret, frame = video_capture.read()
    if not ret:
        break

    frame_count += 1

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.2, 5)

    for (x, y, w, h) in faces:

        face_id = (x, y, w, h)

        if frame_count % process_every_n_frames == 0:

            face_crop = frame[y:y+h, x:x+w]
            face_crop = cv2.resize(face_crop, (160, 160))

            try:
                result = DeepFace.represent(
                    img_path=face_crop,
                    model_name="Facenet",
                    detector_backend="skip",
                    enforce_detection=False
                )

                face_embedding = result[0]["embedding"]

                best_similarity = 0
                best_match_name = "Unknown"

                for known_embedding, person_name in zip(known_encodings, known_names):
                    similarity = cosine_similarity(face_embedding, known_embedding)

                    if similarity > best_similarity:
                        best_similarity = similarity
                        best_match_name = person_name

                if best_similarity > 0.80:
                    name = best_match_name
                    confidence = int(best_similarity * 100)
                else:
                    name = "Unknown"
                    confidence = 0

                face_results[face_id] = (name, confidence)

            except:
                face_results[face_id] = ("Unknown", 0)

        name, confidence = face_results.get(face_id, ("Unknown", 0))

        # Draw
        if name != "Unknown":
            color = (0, 255, 0)
            text = f"{name} ({confidence}%)"
        else:
            color = (0, 0, 255)
            text = "Unknown"

        cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
        cv2.putText(frame, text, (x, y-10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

        # 🔥 MARK PRESENT
        if name != "Unknown" and name not in present_today:

            now = datetime.now()
            time_string = now.strftime("%H:%M:%S")

            # DB
            cursor.execute("""
                INSERT INTO attendance (name, date, time, status)
                VALUES (?, ?, ?, ?)
            """, (name, today_date, time_string, "Present"))
            conn.commit()

            # Excel
            wb = load_workbook(file_name)
            ws = wb.active
            ws.append([name, today_date, time_string, "Present"])
            wb.save(file_name)

            present_today.add(name)

            print(f"{name} marked Present")

    cv2.imshow("Face Recognition Attendance", frame)

    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

# 🔥 MARK ABSENT STUDENTS
absent_students = all_students - present_today

if absent_students:
    wb = load_workbook(file_name)
    ws = wb.active

    for student in absent_students:

        cursor.execute("""
            INSERT INTO attendance (name, date, time, status)
            VALUES (?, ?, ?, ?)
        """, (student, today_date, "-", "Absent"))
        conn.commit()

        ws.append([student, today_date, "-", "Absent"])

        print(f"{student} marked Absent")

    wb.save(file_name)

# ========== CLEANUP ==========
video_capture.release()
conn.close()
cv2.destroyAllWindows()